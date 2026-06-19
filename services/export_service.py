"""
export_service.py — Xuất báo cáo dự báo doanh thu ra PDF / Excel.
Dùng chung cho 2 phạm vi: tổng doanh thu (scope="total") và theo từng
sản phẩm (scope="product"). Input là đúng dict trả về từ
ForecastService.run_forecast() / forecast_product().

LƯU Ý FONT: Helvetica (font mặc định của reportlab) KHÔNG có dấu tiếng Việt
và không có ký hiệu ₫ — nếu dùng sẽ ra ký tự rỗng/ô vuông. Phải nhúng font
TTF hỗ trợ Unicode (DejaVuSans, có sẵn trên Ubuntu tại đường dẫn bên dưới).
"""
import io
import re
import unicodedata
from datetime import datetime
from xml.sax.saxutils import escape

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import matplotlib
matplotlib.use("Agg")  # không cần display, chỉ render ra ảnh
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# FONT UNICODE CHO PDF (tiếng Việt có dấu + ký hiệu ₫)
# ──────────────────────────────────────────────
_FONT_REGULAR = "Helvetica"
_FONT_BOLD = "Helvetica-Bold"
_FONTS_READY = False

_DEJAVU_REGULAR = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
_DEJAVU_BOLD = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"


def _ensure_fonts():
    """Đăng ký font DejaVuSans với reportlab (chỉ chạy 1 lần / process).
    Nếu môi trường không có font này, fallback về Helvetica — PDF vẫn tạo
    được nhưng dấu tiếng Việt và ký hiệu ₫ sẽ không hiển thị đúng."""
    global _FONT_REGULAR, _FONT_BOLD, _FONTS_READY
    if _FONTS_READY:
        return
    try:
        pdfmetrics.registerFont(TTFont("DejaVuSans", _DEJAVU_REGULAR))
        pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", _DEJAVU_BOLD))
        _FONT_REGULAR = "DejaVuSans"
        _FONT_BOLD = "DejaVuSans-Bold"
    except Exception:
        pass
    _FONTS_READY = True


def _fmt_money(v):
    try:
        return f"{v:,.0f} ₫"
    except (TypeError, ValueError):
        return "—"


def _slugify_filename(text: str) -> str:
    """Bỏ dấu tiếng Việt để tên file an toàn trên mọi hệ điều hành.
    Nội dung báo cáo bên trong vẫn giữ đầy đủ dấu — chỉ tên file bị bỏ dấu."""
    norm = unicodedata.normalize("NFKD", text)
    no_diacritics = "".join(c for c in norm if not unicodedata.combining(c))
    no_diacritics = no_diacritics.replace("Đ", "D").replace("đ", "d")
    slug = re.sub(r"[^a-zA-Z0-9_-]+", "_", no_diacritics).strip("_")
    return slug or "bao_cao"


def _model_label(forecast: dict) -> str:
    chosen = forecast.get("model_chosen") or forecast.get("model_type")
    return "Random Forest Regressor" if chosen == "random_forest" else "Linear Regression"


# ──────────────────────────────────────────────
# CHART (dùng chung cho PDF) — vẽ bằng matplotlib, trả PNG dạng BytesIO
# ──────────────────────────────────────────────
def _make_chart_image(actual_labels, actual_values, forecast_labels, forecast_values, title):
    fig, ax = plt.subplots(figsize=(7, 3.1), dpi=150)

    all_labels = list(actual_labels) + list(forecast_labels)
    x_actual = list(range(len(actual_labels)))
    x_forecast = list(range(len(actual_labels) - 1, len(actual_labels) + len(forecast_labels)))
    y_forecast = [actual_values[-1]] + list(forecast_values)

    ax.plot(x_actual, actual_values, marker="o", markersize=4, color="#6366f1", linewidth=2, label="Thực tế")
    ax.plot(x_forecast, y_forecast, marker="o", markersize=4, color="#f97316", linewidth=2, linestyle="--", label="Dự báo")

    ax.set_xticks(range(len(all_labels)))
    ax.set_xticklabels(all_labels, rotation=45, ha="right", fontsize=7)
    ax.set_title(title, fontsize=11)
    ax.legend(fontsize=8, loc="upper left")
    ax.grid(alpha=0.25)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v/1e6:.1f}M"))
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png")
    plt.close(fig)
    buf.seek(0)
    return buf


def _styled_table(rows, col_widths, header_hex="#6366f1"):
    t = Table(rows, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _FONT_REGULAR),
        ("FONTNAME", (0, 0), (-1, 0), _FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_hex)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8faff")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return t


# ──────────────────────────────────────────────
# PDF
# ──────────────────────────────────────────────
def build_pdf_report(scope: str, forecast: dict, product_name: str = None) -> io.BytesIO:
    """scope: 'total' | 'product'. Trả về BytesIO chứa nội dung PDF."""
    _ensure_fonts()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("VNTitle", parent=styles["Title"], fontName=_FONT_BOLD,
                                  alignment=TA_CENTER, fontSize=18, textColor=colors.HexColor("#1e293b"))
    sub_style = ParagraphStyle("VNSub", parent=styles["Normal"], fontName=_FONT_REGULAR,
                                alignment=TA_CENTER, fontSize=12, textColor=colors.HexColor("#6366f1"))
    small_style = ParagraphStyle("VNSmall", parent=styles["Normal"], fontName=_FONT_REGULAR,
                                  fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    h2_style = ParagraphStyle("VNH2", parent=styles["Heading2"], fontName=_FONT_BOLD,
                               fontSize=13, spaceBefore=14, spaceAfter=6, textColor=colors.HexColor("#1e293b"))
    normal_style = ParagraphStyle("VNNormal", parent=styles["Normal"], fontName=_FONT_REGULAR, fontSize=10, leading=14)

    story = []

    if scope == "product":
        story.append(Paragraph("BÁO CÁO DỰ BÁO SẢN PHẨM", title_style))
        story.append(Paragraph(escape(product_name or "Sản phẩm"), sub_style))
    else:
        story.append(Paragraph("BÁO CÁO DỰ BÁO DOANH THU TỔNG", title_style))
        story.append(Paragraph("SalesManager Pro", sub_style))

    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Tạo lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}", small_style))
    story.append(Spacer(1, 14))

    if not forecast.get("has_data"):
        story.append(Paragraph(escape(forecast.get("message", "Không đủ dữ liệu để dự báo.")), normal_style))
        doc.build(story)
        buf.seek(0)
        return buf

    # ── 1. Thông tin mô hình ──
    story.append(Paragraph("1. Thông tin mô hình", h2_style))
    model_label = _model_label(forecast)
    info_rows = [["Thuật toán", model_label]]

    if scope == "total":
        info_rows += [
            ["Số tháng huấn luyện", str(forecast.get("n_train", "-"))],
            ["MAE (sai số tuyệt đối TB)", _fmt_money(forecast.get("mae", 0))],
            ["RMSE", _fmt_money(forecast.get("rmse", 0))],
            ["R² Score", f"{forecast.get('r2', 0):.4f}"],
        ]
        mc = forecast.get("model_comparison")
        if mc:
            better = "Random Forest" if mc["better"] == "random_forest" else "Linear Regression"
            info_rows += [
                ["R² Linear Regression", f"{mc['linear']:.4f}"],
                ["R² Random Forest", f"{mc['random_forest']:.4f}"],
                ["Mô hình tốt hơn", better],
            ]
    else:
        info_rows += [
            ["MAE (sai số tuyệt đối TB)", _fmt_money(forecast.get("mae", 0))],
            ["R² (mô hình đã chọn)", f"{forecast.get('r2', 0):.4f}"],
            ["R² Linear Regression", f"{forecast['r2_linear']:.4f}" if forecast.get("r2_linear") is not None else "—"],
            ["R² Random Forest", f"{forecast['r2_rf']:.4f}" if forecast.get("r2_rf") is not None else "Cần ≥ 4 tháng dữ liệu"],
        ]

    info_table = Table(info_rows, colWidths=[6 * cm, 10.5 * cm])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), _FONT_REGULAR),
        ("FONTNAME", (0, 0), (0, -1), _FONT_BOLD),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f1f5f9")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 12))

    # ── 2. Biểu đồ ──
    chart_buf = _make_chart_image(
        forecast["actual_labels"], forecast["actual_values"],
        forecast["forecast_labels"], forecast["forecast_values"],
        "Doanh thu thực tế vs dự báo",
    )
    chart_w = 16 * cm
    chart_h = chart_w * (3.1 / 7)
    story.append(Image(chart_buf, width=chart_w, height=chart_h))
    story.append(Spacer(1, 12))

    # ── 3. Bảng dữ liệu lịch sử ──
    story.append(Paragraph("2. Dữ liệu lịch sử", h2_style))
    if scope == "product" and forecast.get("actual_qty"):
        rows = [["Tháng", "Doanh thu", "Số lượng"]] + [
            [lbl, _fmt_money(rev), str(int(qty))]
            for lbl, rev, qty in zip(forecast["actual_labels"], forecast["actual_values"], forecast["actual_qty"])
        ]
        col_widths = [5 * cm, 6 * cm, 5.5 * cm]
    else:
        rows = [["Tháng", "Doanh thu"]] + [
            [lbl, _fmt_money(rev)] for lbl, rev in zip(forecast["actual_labels"], forecast["actual_values"])
        ]
        col_widths = [8 * cm, 8.5 * cm]
    story.append(_styled_table(rows, col_widths))
    story.append(Spacer(1, 12))

    # ── 4. Bảng dự báo ──
    story.append(Paragraph("3. Dự báo các tháng tới", h2_style))
    if scope == "product" and forecast.get("forecast_qty"):
        rows = [["Tháng", "Doanh thu dự báo", "Số lượng dự báo"]] + [
            [lbl, _fmt_money(rev), f"{qty:.1f}"]
            for lbl, rev, qty in zip(forecast["forecast_labels"], forecast["forecast_values"], forecast["forecast_qty"])
        ]
        col_widths = [5 * cm, 6 * cm, 5.5 * cm]
    else:
        rows = [["Tháng", "Doanh thu dự báo"]] + [
            [lbl, _fmt_money(rev)] for lbl, rev in zip(forecast["forecast_labels"], forecast["forecast_values"])
        ]
        col_widths = [8 * cm, 8.5 * cm]
    story.append(_styled_table(rows, col_widths, header_hex="#16a34a"))

    story.append(Spacer(1, 18))
    story.append(Paragraph("Báo cáo được tạo tự động bởi hệ thống Dự báo doanh thu AI — SalesManager Pro.", small_style))

    doc.build(story)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────
# EXCEL
# ──────────────────────────────────────────────
def _write_sheet_table(ws, headers, data_rows, header_fill, header_font, formats=None):
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for r, row_data in enumerate(data_rows, start=2):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            if formats and c - 1 < len(formats) and formats[c - 1]:
                cell.number_format = formats[c - 1]
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22


def build_excel_report(scope: str, forecast: dict, product_name: str = None) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = "Tong quan"

    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)

    title = f"BÁO CÁO DỰ BÁO SẢN PHẨM: {product_name}" if scope == "product" else "BÁO CÁO DỰ BÁO DOANH THU TỔNG"
    ws_info["A1"] = title
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A2"] = f"Tạo lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_info["A2"].font = Font(italic=True, color="666666")

    if not forecast.get("has_data"):
        ws_info["A4"] = forecast.get("message", "Không đủ dữ liệu để dự báo.")
        ws_info.column_dimensions["A"].width = 60
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    model_label = _model_label(forecast)
    row = 4
    ws_info.cell(row=row, column=1, value="Thuật toán").font = bold_font
    ws_info.cell(row=row, column=2, value=model_label)
    row += 1

    if scope == "total":
        pairs = [
            ("Số tháng huấn luyện", forecast.get("n_train")),
            ("MAE", forecast.get("mae")),
            ("RMSE", forecast.get("rmse")),
            ("R² Score", forecast.get("r2")),
        ]
        mc = forecast.get("model_comparison")
        if mc:
            pairs += [
                ("R² Linear Regression", mc["linear"]),
                ("R² Random Forest", mc["random_forest"]),
                ("Model tốt hơn", "Random Forest" if mc["better"] == "random_forest" else "Linear Regression"),
            ]
    else:
        pairs = [
            ("MAE", forecast.get("mae")),
            ("R² (mô hình đã chọn)", forecast.get("r2")),
            ("R² Linear Regression", forecast.get("r2_linear")),
            ("R² Random Forest", forecast.get("r2_rf") if forecast.get("r2_rf") is not None else "N/A"),
        ]

    for label, value in pairs:
        ws_info.cell(row=row, column=1, value=label).font = bold_font
        ws_info.cell(row=row, column=2, value=value)
        row += 1

    ws_info.column_dimensions["A"].width = 30
    ws_info.column_dimensions["B"].width = 24

    # ── Sheet: Lich su ──
    ws_hist = wb.create_sheet("Lich su")
    if scope == "product" and forecast.get("actual_qty"):
        headers = ["Tháng", "Doanh thu", "Số lượng"]
        data_rows = list(zip(forecast["actual_labels"], forecast["actual_values"], forecast["actual_qty"]))
        formats = [None, "#,##0", "#,##0"]
    else:
        headers = ["Tháng", "Doanh thu"]
        data_rows = list(zip(forecast["actual_labels"], forecast["actual_values"]))
        formats = [None, "#,##0"]
    _write_sheet_table(ws_hist, headers, data_rows, header_fill, header_font, formats)

    # ── Sheet: Du bao ──
    ws_fc = wb.create_sheet("Du bao")
    if scope == "product" and forecast.get("forecast_qty"):
        headers = ["Tháng", "Doanh thu dự báo", "Số lượng dự báo"]
        data_rows = list(zip(forecast["forecast_labels"], forecast["forecast_values"], forecast["forecast_qty"]))
        formats = [None, "#,##0", "#,##0.0"]
    else:
        headers = ["Tháng", "Doanh thu dự báo"]
        data_rows = list(zip(forecast["forecast_labels"], forecast["forecast_values"]))
        formats = [None, "#,##0"]
    _write_sheet_table(ws_fc, headers, data_rows, header_fill, header_font, formats)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf